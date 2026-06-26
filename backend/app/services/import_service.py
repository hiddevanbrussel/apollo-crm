"""Parse uploaded spreadsheets (xlsx / csv) into normalized row dicts.

Rows are returned keyed by their *original* (whitespace-trimmed) header so the
caller can preserve every column. Use :func:`canonical_field` or
:func:`contact_canonical_field` to map headers onto known CRM fields.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook

# Company import: header aliases -> canonical CRM field.
HEADER_ALIASES: dict[str, str] = {
    "customer_name": "customer_name",
    "customer name": "customer_name",
    "company": "customer_name",
    "company_name": "customer_name",
    "name": "customer_name",
    "bedrijf": "customer_name",
    "bedrijfsnaam": "customer_name",
    "country": "country",
    "land": "country",
    "domain": "domain",
    "website": "domain",
    "website_url": "domain",
}

# Contact import: header aliases -> canonical contact field.
# ``customer_name`` links the person to an existing company (by name).
CONTACT_HEADER_ALIASES: dict[str, str] = {
    "customer_name": "customer_name",
    "customer name": "customer_name",
    "company": "customer_name",
    "company_name": "customer_name",
    "bedrijf": "customer_name",
    "bedrijfsnaam": "customer_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "voornaam": "first_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "achternaam": "last_name",
    "full_name": "full_name",
    "fullname": "full_name",
    "name": "full_name",
    "naam": "full_name",
    "contact_name": "full_name",
    "email": "email",
    "e-mail": "email",
    "e_mail": "email",
    "phone": "phone",
    "telefoon": "phone",
    "mobile": "phone",
    "title": "title",
    "job_title": "title",
    "functie": "title",
    "linkedin_url": "linkedin_url",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "city": "city",
    "stad": "city",
    "country": "country",
    "land": "country",
    "seniority": "seniority",
    "department": "department",
    "afdeling": "department",
}

# Spreadsheet placeholders that mean "no domain" (Dutch/English).
DOMAIN_PLACEHOLDERS = frozenset(
    {"", "niet", "geen", "n/a", "na", "none", "unknown", "-", "—", "null", "?"}
)


class ImportParseError(Exception):
    pass


def canonical_field(header: object) -> str | None:
    """Return the canonical company field for a header, or None if it is custom."""
    if header is None:
        return None
    return HEADER_ALIASES.get(str(header).strip().lower())


def contact_canonical_field(header: object) -> str | None:
    """Return the canonical contact field for a header, or None if it is custom."""
    if header is None:
        return None
    return CONTACT_HEADER_ALIASES.get(str(header).strip().lower())


def normalize_domain(value: str | None) -> str | None:
    """Turn a spreadsheet domain/website cell into a bare hostname, or None if empty/placeholder."""
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    lowered = raw.lower()
    if lowered in DOMAIN_PLACEHOLDERS:
        return None
    candidate = raw
    if "://" in candidate or candidate.startswith("www."):
        parsed = urlparse(candidate if "://" in candidate else f"https://{candidate}")
        candidate = parsed.netloc or parsed.path.split("/")[0]
    candidate = candidate.strip().lower().removeprefix("www.")
    if not candidate or candidate in DOMAIN_PLACEHOLDERS or "@" in candidate or " " in candidate:
        return None
    return candidate


def _has_name_column(headers) -> bool:
    return any(canonical_field(h) == "customer_name" for h in headers)


def _has_contact_columns(headers) -> bool:
    has_company = any(contact_canonical_field(h) == "customer_name" for h in headers)
    has_identity = any(
        contact_canonical_field(h) in {"first_name", "last_name", "full_name", "email"} for h in headers
    )
    return has_company and has_identity


def _parse_xlsx(content: bytes, *, validate_headers: Callable[[list], bool], missing_msg: str) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ImportParseError(f"Could not read Excel file: {exc}") from exc
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h).strip() if h not in (None, "") else None for h in header_row]
    if not validate_headers(headers):
        raise ImportParseError(missing_msg)

    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if header is None or idx >= len(raw):
                continue
            value = raw[idx]
            record[header] = str(value).strip() if value not in (None, "") else ""
        if any(v for v in record.values()):
            rows.append(record)
    return rows


def _parse_csv(content: bytes, *, validate_headers: Callable[[list], bool], missing_msg: str) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_list = list(reader)
    if not rows_list:
        return []

    headers = [h.strip() if h and h.strip() else None for h in rows_list[0]]
    if not validate_headers(headers):
        raise ImportParseError(missing_msg)

    rows: list[dict[str, str]] = []
    for raw in rows_list[1:]:
        record: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if header is None or idx >= len(raw):
                continue
            value = raw[idx]
            record[header] = value.strip() if value else ""
        if any(v for v in record.values()):
            rows.append(record)
    return rows


_COMPANY_MISSING = (
    "Required column 'customer_name' not found. Expected at least: customer_name."
)
_CONTACT_MISSING = (
    "Required columns not found. Expected customer_name (company) plus at least one of: "
    "first_name, last_name, full_name, email."
)


def _parse_spreadsheet(
    filename: str,
    content: bytes,
    *,
    validate_headers: Callable[[list], bool],
    missing_msg: str,
) -> list[dict[str, str]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return _parse_csv(content, validate_headers=validate_headers, missing_msg=missing_msg)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _parse_xlsx(content, validate_headers=validate_headers, missing_msg=missing_msg)
    try:
        return _parse_xlsx(content, validate_headers=validate_headers, missing_msg=missing_msg)
    except ImportParseError:
        return _parse_csv(content, validate_headers=validate_headers, missing_msg=missing_msg)


def parse_spreadsheet(filename: str, content: bytes) -> list[dict[str, str]]:
    """Return company-import rows as {original_header: value} dicts."""
    return _parse_spreadsheet(
        filename,
        content,
        validate_headers=_has_name_column,
        missing_msg=_COMPANY_MISSING,
    )


def parse_contact_spreadsheet(filename: str, content: bytes) -> list[dict[str, str]]:
    """Return contact-import rows as {original_header: value} dicts."""
    return _parse_spreadsheet(
        filename,
        content,
        validate_headers=_has_contact_columns,
        missing_msg=_CONTACT_MISSING,
    )


CONTACT_FIELDS = frozenset(
    {
        "customer_name",
        "first_name",
        "last_name",
        "full_name",
        "email",
        "phone",
        "title",
        "linkedin_url",
        "city",
        "country",
        "seniority",
        "department",
    }
)


def extract_contact_row(row: dict[str, str]) -> tuple[dict[str, str | None], dict[str, str]]:
    """Split a raw spreadsheet row into contact fields and leftover extra columns."""
    fields: dict[str, str | None] = {k: None for k in CONTACT_FIELDS}
    extra: dict[str, str] = {}
    for header, value in row.items():
        field = contact_canonical_field(header)
        raw = (value or "").strip()
        if field in CONTACT_FIELDS:
            if raw:
                fields[field] = raw
        elif raw:
            extra[header] = raw
    return fields, extra
